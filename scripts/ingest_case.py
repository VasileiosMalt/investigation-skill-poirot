"""
ingest_case.py — Phase 1: Case directory scanner and text evidence extractor.

Usage:
    python ingest_case.py --case-dir /path/to/case --output-dir /path/to/output

Outputs:
    case_manifest.json         — Full file inventory with modality classification
    case_knowledge_base.md     — Structured CKB: entities, timeline, patterns
    phase1_notes.md            — Raw analyst notes before filtering
"""

import os
import sys
import json
import hashlib
import argparse
import mimetypes
from datetime import datetime
from pathlib import Path
from typing import Optional

# ── Optional imports (install as needed) ──────────────────────────────────────
try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    import extract_msg
    HAS_MSG = True
except ImportError:
    HAS_MSG = False

# ── Modality classification ────────────────────────────────────────────────────

TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".log", ".csv", ".tsv",
    ".json", ".yaml", ".yml", ".xml", ".html", ".htm",
    ".eml", ".rtf"
}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".doc", ".odt", ".ods", ".msg"}
DATA_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv", ".json", ".xml", ".yaml", ".yml"}
IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".webp", ".gif", ".tiff", ".tif",
    ".bmp", ".heic", ".heif", ".raw", ".cr2", ".cr3", ".nef",
    ".nrw", ".arw", ".dng", ".svg", ".avif"
}
AUDIO_EXTENSIONS = {
    ".mp3", ".wav", ".flac", ".aac", ".m4a", ".ogg", ".opus",
    ".wma", ".aiff", ".aif", ".amr"
}
VIDEO_EXTENSIONS = {
    ".mp4", ".mov", ".avi", ".mkv", ".wmv", ".webm", ".flv",
    ".mpeg", ".mpg", ".ts", ".m2ts"
}
SKIP_PATTERNS = {
    "__macosx", ".ds_store", "thumbs.db", ".thumbnails",
    "desktop.ini", ".git", ".svn", "node_modules"
}


def classify_file(path: Path) -> dict:
    """Classify a file by modality and subtype."""
    ext = path.suffix.lower()
    name_lower = path.name.lower()

    # Skip system/junk files
    for skip in SKIP_PATTERNS:
        if skip in str(path).lower():
            return None

    if ext in IMAGE_EXTENSIONS:
        modality = "image"
        subtype = "raw_photo" if ext in {".raw", ".cr2", ".cr3", ".nef", ".nrw", ".arw", ".dng"} else "photograph"
    elif ext in AUDIO_EXTENSIONS:
        modality = "audio"
        subtype = "audio"
    elif ext in VIDEO_EXTENSIONS:
        modality = "video"
        subtype = "video"
    elif ext == ".pdf":
        modality = "document"
        subtype = "pdf"  # Will be refined during reading
    elif ext in {".docx", ".doc"}:
        modality = "document"
        subtype = "word"
    elif ext in {".xlsx", ".xls"}:
        modality = "data"
        subtype = "spreadsheet"
    elif ext in DATA_EXTENSIONS:
        modality = "data"
        subtype = "structured"
    elif ext in TEXT_EXTENSIONS or ext == ".msg" or ext == ".eml":
        modality = "text"
        subtype = "email" if ext in {".eml", ".msg"} else "raw_text"
    else:
        # Attempt MIME detection
        mime, _ = mimetypes.guess_type(str(path))
        if mime:
            if mime.startswith("image/"): modality, subtype = "image", "photograph"
            elif mime.startswith("audio/"): modality, subtype = "audio", "audio"
            elif mime.startswith("video/"): modality, subtype = "video", "video"
            elif mime.startswith("text/"): modality, subtype = "text", "raw_text"
            else: modality, subtype = "unknown", "binary"
        else:
            modality, subtype = "unknown", "binary"

    try:
        stat = path.stat()
        size_bytes = stat.st_size
        mtime = datetime.fromtimestamp(stat.st_mtime).isoformat()
        ctime = datetime.fromtimestamp(stat.st_ctime).isoformat()
    except OSError:
        size_bytes, mtime, ctime = 0, None, None

    # File hash (MD5 for deduplication — fast enough for investigation use)
    try:
        md5 = hashlib.md5(path.read_bytes()).hexdigest()
    except (OSError, MemoryError):
        md5 = None

    return {
        "path": str(path),
        "name": path.name,
        "extension": ext,
        "modality": modality,
        "subtype": subtype,
        "size_bytes": size_bytes,
        "modified": mtime,
        "created": ctime,
        "md5": md5,
        "analysis_lane": f"phase{'3' if modality == 'image' else '4' if modality in ('audio','video') else '1'}",
    }


def scan_case_directory(case_dir: Path) -> list[dict]:
    """Recursively scan the case directory and classify all files."""
    files = []
    seen_hashes = {}

    for fpath in sorted(case_dir.rglob("*")):
        if not fpath.is_file():
            continue

        entry = classify_file(fpath)
        if entry is None:
            continue

        # Flag duplicates
        if entry["md5"] and entry["md5"] in seen_hashes:
            entry["duplicate_of"] = seen_hashes[entry["md5"]]
        elif entry["md5"]:
            seen_hashes[entry["md5"]] = str(fpath)

        # Relative path from case root
        try:
            entry["relative_path"] = str(fpath.relative_to(case_dir))
        except ValueError:
            entry["relative_path"] = entry["path"]

        files.append(entry)

    return files


# ── Text extraction ────────────────────────────────────────────────────────────

def extract_text_from_file(entry: dict) -> Optional[str]:
    """Extract raw text content from a file based on its type."""
    path = Path(entry["path"])
    ext = entry["extension"]

    try:
        if ext in {".txt", ".md", ".log", ".csv", ".tsv", ".html", ".htm", ".rtf"}:
            return path.read_text(encoding="utf-8", errors="replace")

        elif ext in {".json"}:
            data = json.loads(path.read_bytes())
            return json.dumps(data, indent=2)

        elif ext in {".yaml", ".yml"}:
            return path.read_text(encoding="utf-8", errors="replace")

        elif ext == ".xml":
            return path.read_text(encoding="utf-8", errors="replace")

        elif ext == ".pdf" and HAS_PDF:
            text_pages = []
            with pdfplumber.open(path) as pdf:
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if text:
                        text_pages.append(f"[Page {i+1}]\n{text}")
            if text_pages:
                entry["subtype"] = "pdf_text"
                return "\n\n".join(text_pages)
            else:
                entry["subtype"] = "pdf_scanned"
                entry["analysis_lane"] = "phase3"
                return None

        elif ext in {".docx"} and HAS_DOCX:
            doc = DocxDocument(path)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

        elif ext in {".xlsx", ".xls"} and HAS_XLSX:
            wb = openpyxl.load_workbook(path, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"[Sheet: {sheet_name}]")
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip():
                        lines.append(row_text)
            return "\n".join(lines)

        elif ext == ".msg" and HAS_MSG:
            msg = extract_msg.openMsg(str(path))
            parts = []
            if msg.subject: parts.append(f"Subject: {msg.subject}")
            if msg.sender: parts.append(f"From: {msg.sender}")
            if msg.to: parts.append(f"To: {msg.to}")
            if msg.date: parts.append(f"Date: {msg.date}")
            if msg.body: parts.append(f"\nBody:\n{msg.body}")
            return "\n".join(parts)

        elif ext == ".eml":
            import email
            msg = email.message_from_bytes(path.read_bytes())
            parts = [f"Subject: {msg.get('Subject', '')}",
                     f"From: {msg.get('From', '')}",
                     f"To: {msg.get('To', '')}",
                     f"Date: {msg.get('Date', '')}"]
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    parts.append(f"\nBody:\n{part.get_payload(decode=True).decode('utf-8', errors='replace')}")
                    break
            return "\n".join(parts)

    except Exception as e:
        entry["extraction_error"] = str(e)

    return None


# ── Priority detection (case description files) ──────────────────────────────

CASE_DESCRIPTION_NAMES = {
    "case", "description", "brief", "readme", "summary", "overview",
    "incident", "report", "complaint", "case_file", "case_notes"
}

def is_case_description(name: str) -> bool:
    stem = Path(name).stem.lower().replace("-", "_").replace(" ", "_")
    return any(kw in stem for kw in CASE_DESCRIPTION_NAMES)


def prioritise_text_files(files: list[dict]) -> list[dict]:
    """Sort text/document files: case description first, then by path."""
    def sort_key(f):
        if f["modality"] in ("text", "document", "data"):
            return (0 if is_case_description(f["name"]) else 1, f["relative_path"])
        return (2, f["relative_path"])
    return sorted(files, key=sort_key)


# ── Main ───────────────────────────────────────────────────────────────────────

def build_manifest(case_dir: Path, files: list[dict]) -> dict:
    counts = {}
    for f in files:
        counts[f["modality"]] = counts.get(f["modality"], 0) + 1

    return {
        "case_root": str(case_dir),
        "scanned_at": datetime.utcnow().isoformat() + "Z",
        "total_files": len(files),
        "modality_counts": counts,
        "files": files
    }


def build_phase1_notes(case_dir: Path, files: list[dict], texts: dict) -> str:
    lines = [
        f"# Phase 1 Notes — {case_dir.name}",
        f"Generated: {datetime.utcnow().isoformat()}Z\n",
        "## File Inventory Summary",
        f"Total files: {len(files)}"
    ]

    counts = {}
    for f in files:
        counts[f["modality"]] = counts.get(f["modality"], 0) + 1
    for mod, cnt in sorted(counts.items()):
        lines.append(f"  - {mod}: {cnt}")

    lines.append("\n## Text Content Extracted\n")
    for fname, text in texts.items():
        lines.append(f"### {fname}")
        lines.append(text[:3000] + ("\n[... truncated ...]" if len(text) > 3000 else ""))
        lines.append("")

    return "\n".join(lines)


def generate_ckb_prompt(case_dir_name: str, all_text: str) -> str:
    """
    Returns the LLM prompt used to generate the Case Knowledge Base.
    The caller should send this to the preferred LLM and write the response
    to case_knowledge_base.md.
    """
    return f"""You are a meticulous forensic analyst building a Case Knowledge Base for an investigation.

CASE DIRECTORY: {case_dir_name}

EVIDENCE TEXT (all readable text files concatenated):
---
{all_text[:40000]}
---

Your task: Produce a structured Case Knowledge Base (CKB) in Markdown with EXACTLY these sections:

## Case Summary
[2-4 sentences: what happened, who is involved, what question needs to be answered]

## Known Entities
### People
| Name | Role | Key Facts | First Mentioned In |
|---|---|---|---|

### Places
| Name | Relevance | Mentioned In |
|---|---|---|

### Timeline
| Date/Time | Event | Source | Confidence |
|---|---|---|---|

### Key Objects / Items
| Item | Relevance | Mentioned In |
|---|---|---|

## Contradictions & Inconsistencies
[List ONLY confirmed contradictions with specific source citations. If none found, write "None detected."]

## Information Gaps
[List information that is conspicuously absent given the case context]

## Pattern Observations
[List ONLY patterns supported by 2+ data points. Each must cite evidence sources. If none, write "None detected."]

Rules:
- Be precise and factual. No speculation at this stage.
- Only include entities, dates, and facts that appear in the evidence text.
- Flag [UNCLEAR] for items that appear referenced but without enough context to be certain.
- Quality over quantity — a short accurate CKB is better than a long inaccurate one.
"""


def main():
    parser = argparse.ArgumentParser(description="Poirot Phase 1 — Case Ingestion")
    parser.add_argument("--case-dir", required=True, help="Path to the case directory")
    parser.add_argument("--output-dir", required=True, help="Where to write output files")
    parser.add_argument("--print-ckb-prompt", action="store_true",
                        help="Print the LLM prompt for CKB generation and exit")
    args = parser.parse_args()

    case_dir = Path(args.case_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not case_dir.exists():
        print(f"ERROR: Case directory not found: {case_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"[Poirot Phase 1] Scanning: {case_dir}")
    files = scan_case_directory(case_dir)
    print(f"  Found {len(files)} files")

    # Extract text from all readable files
    text_content = {}
    ordered = prioritise_text_files(files)
    for entry in ordered:
        if entry["modality"] in ("text", "document", "data"):
            text = extract_text_from_file(entry)
            if text:
                text_content[entry["relative_path"]] = text
                print(f"  [TEXT] {entry['relative_path']} ({len(text)} chars)")

    # Write case_manifest.json
    manifest = build_manifest(case_dir, files)
    manifest_path = output_dir / "case_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, default=str))
    print(f"[Poirot Phase 1] Written: {manifest_path}")

    # Write phase1_notes.md
    notes = build_phase1_notes(case_dir, files, text_content)
    notes_path = output_dir / "phase1_notes.md"
    notes_path.write_text(notes, encoding="utf-8")
    print(f"[Poirot Phase 1] Written: {notes_path}")

    # Write or print CKB prompt
    all_text = "\n\n---\n\n".join(
        f"[FILE: {fname}]\n{text}" for fname, text in text_content.items()
    )
    ckb_prompt = generate_ckb_prompt(case_dir.name, all_text)

    if args.print_ckb_prompt:
        print("\n" + "="*60)
        print("CKB GENERATION PROMPT (send to LLM):")
        print("="*60)
        print(ckb_prompt)
    else:
        prompt_path = output_dir / "ckb_prompt.txt"
        prompt_path.write_text(ckb_prompt, encoding="utf-8")
        print(f"[Poirot Phase 1] CKB prompt written to: {prompt_path}")
        print("  → Send ckb_prompt.txt to your preferred LLM and save the response as case_knowledge_base.md")

    print(f"\n[Poirot Phase 1] Complete. Outputs in: {output_dir}")
    print(f"  Next: run classify_evidence.py with the manifest")


if __name__ == "__main__":
    main()
